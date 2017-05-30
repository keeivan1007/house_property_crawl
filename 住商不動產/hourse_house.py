"""
purpose:爬取東森房屋菜單頁面資料

input:selenium完畢，已回傳的未整理資料結構

output:整理好的zip

※支援型方法

會使用zip的原因是因為剖析出的資訊都是一排的(一組list)

在下個程式要輸入至csv無法用換欄位去存取，只能一行一行，

在這個前提下資料勢必得包裝成，多個屬性的資訊，轉成多筆資料的資訊



資料:

('王曉明','呵呵呵','狂笑狂')

(27,55,11)

→(('王曉明',27),('呵呵呵',55),('狂笑狂',11))

"""

def crawal_soup(soup):
    
    import re
    www = soup.find_all(class_ = 'InfoWrap')#找出當頁個別資訊

    #記得要先置空
    title_list = []
    address = []
    build_pings_list = []
    price_list = []
    class_list = []
    pattern_list = []
    land_list = []
    floor_list = []
    ID_list = []
    url_list = []


    for i in www:
        total = i.text.split('\n') #去行
        filter_list = list(filter(None,total)) #過濾掉list裡面有空的值，並list化

        regex = re.compile("低於行情")
        aaa = list(filter(regex.search,filter_list))
        if aaa != []:
            filter_list.remove(aaa[0])



        #放置區

        #標題  #地址
        title_list.append(filter_list[0])
        address.append(filter_list[1])


        #建坪
        regex = re.compile("建坪：")
        build_pings = list(filter(regex.search,filter_list))
        if build_pings != []:
            build_pings_list.append(build_pings[0].split('：')[1].split('坪')[0])
        else:
            build_pings_list.append('')

        #價格
        regex = re.compile("[0-9]+萬         [0-9]+人")
        price = list(filter(regex.search,filter_list))
        if price != []:
            price_list.append(price[0].split('         ')[0])
        else:
            price_list.append('')

        #類型 #格局
        regex = re.compile("│")
        house_cp = list(filter(regex.search,filter_list))
        if house_cp != []:
            class_list.append(house_cp[0].split('│')[0])
            pattern_list.append(house_cp[0].split('│')[1].replace(' ',''))
        else:
            class_list.append('')
            pattern_list.append('')

        #地坪
        regex = re.compile("地坪：")
        land = list(filter(regex.search,filter_list))
        if land != []:
            land_list.append(land[0].split('：')[1].split('坪')[0])
        else:
            land_list.append('')

        #樓層
        regex = re.compile("樓層:")
        floor = list(filter(regex.search,filter_list))
        if floor[0] == '/':  #搞定沒填寫樓層的部分
            floor_list.append('')
        elif floor != []:
            over_text = ''.join(floor[0].split(' ')[:-1])
            floor_list.append(over_text.split(':')[1])
        else:
            floor_list.append('')        



        #找出 ID 與 網址

    sss = soup.find_all(href=re.compile("mapstreetwrap")) #透過 mapstreetwrap 找出案件ID
    for a in sss:  #需要取出中間那個值
        hhh = 'http://www.hbhousing.com.tw' + a['href'].split('#')[0]
        url_list.append(hhh)
        eee = hhh.split('sn=')[1]
        ID_list.append(eee)



    total = zip(ID_list,title_list,class_list,pattern_list,price_list,build_pings_list,land_list,floor_list,url_list,address)

    return total



"""
propose:把塞選好的資料轉成xlsx存放

input:整理好的zip檔案 ；要更新的檔案xlsx

output:整理好的xlsx存放

1.檔案請放同個路徑

2.每次存取為20，住商不動產專用

3.欄位格式、sheet要固定


範例：

updata_xlsx_file = '測試資料.xlsx'


save_excel_Eastern(total_info,updata_xlsx_file)

"""


def save_excel_House(total_info,updata_xlsx_file):
    
    from openpyxl.reader.excel import load_workbook  #讀取用
    from openpyxl.writer.excel import ExcelWriter  #存檔用

    wb=load_workbook(updata_xlsx_file)
    ws = wb.get_sheet_by_name("工作表1")
    row_counts = ws.max_row #取得行列數 

    range_string = "A{}:J{}".format(row_counts+1,row_counts+20) #一共20筆資料要存放
    data = ws.iter_rows(range_string) #取出空的欄位

    '''前置區 把抓到資料用城可讀取'''

    abc_list = [] #給存放的list 才可以到 row
    for i in total_info:
        abc_list.append(i)


    """運算區 資料塞進到excel"""


    row = 0 #從這 抓進去的資料從第0行開始
    for a in data:
        column = 0 #從這 抓進去的資列從第0欄開始
        for cell in a:
            cell.value = abc_list[row][column]
            column = column + 1
        row = row + 1

    ew = ExcelWriter(workbook = wb)  #新建立一個 excelWriter 
    ew.save(filename=updata_xlsx_file)