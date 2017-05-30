"""
purpose:爬取東森房屋菜單頁面資料

input:selenium完畢，已回傳的未整理資料結構

output:整理好的zip

※支援型方法

會使用zip的原因是因為剖析出的資訊都是一排的(一組list)

在下個程式要輸入至csv無法用換欄位去存取，只能一行一行，

在這個前提下資料勢必得包裝成，多個屬性的資訊，轉成多筆資料的資訊



資料:

('王曉明','呵呵呵','狂笑狂')

(27,55,11)

→(('王曉明',27),('呵呵呵',55),('狂笑狂',11))

"""

def crawal_soup(soup):
    
    #標題、地址
    from bs4 import BeautifulSoup
    import re
    text_b = soup.find_all('h3')[:10]  #頁數前十個才是要的資訊 後面不是
    title_list = []
    address_list = []
    for a in text_b:

        title_address = a.text.replace('\n','').replace(' ','').split('　')
        title_list.append(title_address[0])
        address_list.append(title_address[1])

    #價格
    text_b = soup.find_all('div',{'class': 'price'})
    price_list = []
    for a in text_b:
        price_list.append(str(a.text)) #要把它轉成文字

    #代碼跟網址
    text_a = soup.find_all('a',{'class': 'obj_item_photo'})
    ID_list = []
    url_list = []
    for a in text_a:
        ID_list.append(a['href'].split('-')[1])
        url_list.append(a['href'])
        
    #類型、格局
    tails = soup.find_all('ul',class_='obj_detail')
    class_list = [] #類型
    pattern_list = [] #格局
    for a in tails:
        total_tails = a.text.split('\n')
        class_list.append(total_tails[1])
        pattern_list.append(total_tails[2])

        
        
    #坪數跟屋齡
    aaa = soup.find_all(string=re.compile("坪"))
    pings_list = aaa[-10:] #坪數

        
        
    total = zip(ID_list,title_list,class_list,pattern_list,price_list,pings_list,url_list,address_list) #最後彙整成zip型式(緣由請看上)
    
    return total



"""
propose:把塞選好的資料轉成xlsx存放

input:整理好的zip檔案 ，list順序 ID→title→price→url→address ；要更新的檔案xlsx

output:整理好的xlsx存放

1.檔案請放同個路徑

2.每次存取為10，東森購物專用

3.欄位格式、sheet要固定


範例：

updata_xlsx_file = '測試資料.xlsx'


save_excel_Eastern(total_info,updata_xlsx_file)

"""


def save_excel_Eastern(total_info,updata_xlsx_file):
    
    from openpyxl.reader.excel import load_workbook  #讀取用
    from openpyxl.writer.excel import ExcelWriter  #存檔用

    wb=load_workbook(updata_xlsx_file)
    ws = wb.get_sheet_by_name("工作表1")
    row_counts = ws.max_row #取得行列數 

    range_string = "A{}:H{}".format(row_counts+1,row_counts+10) #一共三筆資料要存放
    data = ws.iter_rows(range_string) #取出空的欄位

    '''前置區 把抓到資料用城可讀取'''

    abc_list = [] #給存放的list 才可以到 row
    for i in total_info:
        abc_list.append(i)


    """運算區 資料塞進到excel"""


    row = 0 #從這 抓進去的資料從第0行開始
    for a in data:
        column = 0 #從這 抓進去的資列從第0欄開始
        for cell in a:
            cell.value = abc_list[row][column]
            column = column + 1
        row = row + 1

    ew = ExcelWriter(workbook = wb)  #新建立一個 excelWriter 
    ew.save(filename=updata_xlsx_file)